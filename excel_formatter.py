import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


#格式渲染
#居中对齐
alignment1 = Alignment(horizontal = "center",vertical = "center")
#字体1：雅黑10加粗
font1 = Font(name = "微软雅黑",size = 10,bold = True)
#字体2：雅黑10红色
font2 = Font(name = "微软雅黑",size = 10,color = "FF0000")
#字体3：雅黑10
font3 = Font(name = "微软雅黑",size = 10)
#字体4：雅黑10红色加粗
font4 = Font(name = "微软雅黑",size = 10,bold = True,color = "FF0000")
#淡蓝40%
fill1 = PatternFill("solid",fgColor = "9BC2E6")
#淡蓝60%
fill2 = PatternFill("solid",fgColor = "BDD7EE")
#边框：上下
border1 = Border(top = Side(border_style = "medium",color = "000000"),bottom = Side(border_style = "medium",color = "000000"))
#边框：上下右
border2 = Border(top = Side(border_style = "medium",color = "000000"),bottom = Side(border_style = "medium",color = "000000"),\
                 right = Side(border_style = "medium",color = "000000"))
#边框：右
border3 = Border(right = Side(border_style = "medium",color = "000000"))
#边框：下
border4 = Border(bottom = Side(border_style = "medium",color = "000000"))
#边框：右下
border5 = Border(right = Side(border_style = "medium",color = "000000"),bottom = Side(border_style = "medium",color = "000000"))
#边框：上
border6 = Border(top = Side(border_style = "medium",color = "000000"))
#边框：右上
border7 = Border(top = Side(border_style = "medium",color = "000000"),right = Side(border_style = "medium",color = "000000"))
#边框：细下
border8 = Border(bottom = Side(border_style = "thin",color = "000000"))
#边框：细下中右
border9 = Border(bottom = Side(border_style = "thin",color = "000000"),right = Side(border_style = "medium",color = "000000"))
#边框：细上
border10 = Border(top = Side(border_style = "thin",color = "000000"))
#边框：细上中右
border11 = Border(top = Side(border_style = "thin",color = "000000"),right = Side(border_style = "medium",color = "000000"))
def xls_formatter_cinema(wb,df,field):
    wb.guess_types = True
    sheet_name = "%s数据" % sheet_name_dict[field]
    ws = wb.create_sheet(sheet_name)
    cols_list = df.columns.tolist()
    data_list = np.array(df).tolist()
    #市占部分
    market_percent_list = []
    if field == "bo":
        market_percent_list = data_list[-4:]
        data_list = data_list[:-4]
    
    m = len(data_list)
    n = len(cols_list)
    k = len(market_percent_list)
    #写入首行
    ws.append(cols_list)
    ws.row_dimensions[1].height = 20
    for i in range(1,n + 1):
        cell = ws.cell(row = 1,column = i)
        cell.alignment = alignment1
        cell.font = font1
        cell.fill = fill1
        cell.border = border1
        if i == n:
            cell.border = border2
        if field in ["bo","people","occupancy","avg_price"] and i == 4:
            cell.border = border2
        if field == "consession" and i == 1:
            cell.border = border2
    
    #写入数据
    for each_data in data_list:
        ws.append(each_data)
    #按字段区分两类进行渲染
    if field in ["bo","people","occupancy","avg_price"]:
        for i in range(2,m + 2):
            ws.row_dimensions[i].height = 15
            for j in range(1,n + 1):
                cell = ws.cell(row = i,column = j)
                cell.alignment = alignment1
                if j ==4 or j == n:
                    cell.border = border3
                if i == 2:
                    cell.font = font2
                elif i > 2:
                    cell.font = font3
                if j <= 4:
                    cell.fill = fill2
                else:
                    cell.number_format = "0.00"
        #添加底部边框
        for j in range(1,n + 1):
            cell = ws.cell(row = m + 1,column = j)
            cell.border = border4
            if j == 4 or j == n:
                cell.border = border5
                
        #若有市占数据
        if k != 0:
            for i in range(0,k):
                ws.append(market_percent_list[i])
                ws.row_dimensions[m + 2 + i].height = 15
                for j in range(1,n + 1):
                    cell = ws.cell(row = m + 2 + i,column = j)
                    cell.alignment = alignment1
                    cell.font = font4
                    if j >4:
                        cell.number_format = "0.00%"
                    if j == n:
                        cell.border = border3
            for j in range(1,n + 1):
                cell = ws.cell(row = m + 1 + k,column = j)
                cell.border = border4
                if j == n:
                    cell.border = border5
        
        #第2-8列调整列宽
        ws.column_dimensions[get_column_letter(2)].width = 22
        ws.column_dimensions[get_column_letter(3)].width = 18
        ws.column_dimensions[get_column_letter(4)].width = 10
        for j in range(5,8):
            ws.column_dimensions[get_column_letter(j)].width = 12
        #冻结窗格
        ws.freeze_panes = ws["E2"]
        
    elif field == "consession":
        for i in range(2,m + 2):
            ws.row_dimensions[i].height = 15
            for j in range(1,n + 1):
                cell = ws.cell(row = i,column = j)
                cell.alignment = alignment1
                cell.font = font3
                cell.number_format = "0.00"
                if j == 1:
                    cell.fill = fill2
                    cell.border = border3
                if j == n:
                    cell.border = border3
                    
        for j in range(1,n + 1):
            cell = ws.cell(row = m + 1,column = j)
            cell.border = border4
            if j == 1 or j == n:
                cell.border = border5
                
        ws.column_dimensions[get_column_letter(1)].width = 15
        ws.freeze_panes = ws["B2"]
                
    return wb

def xls_formatter_city(wb,df,field,cinema_index_list):
    wb.guess_types = True
    sheet_name = "%s数据" % sheet_name_dict[field]
    ws = wb.create_sheet(sheet_name)
    cols_list = df.columns.tolist()
    data_list = np.array(df).tolist()
    #附加部分
    attach_list = []
    if field in ["bo"]:
        attach_list = data_list[-5:]
        data_list = data_list[:-5]
    elif field in ["consession"]:
        attach_list = data_list[-4:]
        data_list = data_list[:-4]
    elif field in ["people","occupancy","avg_price"]:
        attach_list = data_list[-1]
        data_list = data_list[:-1]
        
    m = len(data_list)
    n = len(cols_list)
    k = len(attach_list)
    #写入首行
    ws.append(cols_list)
    ws.row_dimensions[1].height = 20
    for i in range(1,n + 1):
        cell = ws.cell(row = 1,column = i)
        cell.alignment = alignment1
        cell.font = font1
        cell.fill = fill1
        cell.border = border1
        if i == n:
            cell.border = border2
        if field in ["bo","people","occupancy","avg_price"] and i == 4:
            cell.border = border2
        if field == "consession" and i == 3:
            cell.border = border2
            
    #写入数据
    for each_data in data_list:
        ws.append(each_data)
    #按字段区分两类进行渲染
    if field in ["bo","people","occupancy","avg_price"]:
        for i in range(2,m + 2):
            ws.row_dimensions[i].height = 15
            for j in range(1,n + 1):
                cell = ws.cell(row = i,column = j)
                cell.alignment = alignment1
                if j == 4 or j == n:
                    cell.border = border3
                #自营影城加红
                if i - 2 in cinema_index_list:
                    cell.font = font2
                    if i - 2 != 0:
                        cell.border = border10
                        if j == 4 or j == n:
                            cell.border = border11
                elif i - 2 not in cinema_index_list:
                    cell.font = font3
                if j <= 4:
                    cell.fill = fill2
                else:
                    cell.number_format = "0.00"
                  
        #添加底部边框
        for j in range(1,n + 1):
            cell = ws.cell(row = m + 1,column = j)
            cell.border = border4
            if j == 4 or j == n:
                cell.border = border5
                
        #附加的汇总部分
        if k == 5:
            for i in range(0,k):
                ws.append(attach_list[i])
                ws.row_dimensions[m + 2 + i].height = 15
                for j in range(1,n + 1):
                    cell = ws.cell(row = m + 2 + i,column = j)
                    cell.alignment = alignment1
                    cell.font = font4
                    if j == n:
                        cell.border = border3
                    if i == 0:
                        cell.number_format = "0.00"
                    elif i != 0:
                        cell.number_format = "0.00%"
            #添加底部边框
            for j in range(1,n+1):
                cell = ws.cell(row = m + 1 + k,column = j)
                cell.border = border4
                if j == n:
                    cell.border = border5
                
        elif k != 5:
            ws.append(attach_list)
            ws.row_dimensions[m + 2].height = 15
            for j in range(1,n + 1):
                cell = ws.cell(row = m + 2,column = j)
                cell.alignment = alignment1
                cell.font = font4
                cell.border = border4
                cell.number_format = "0.00"
                if j == n:
                    cell.border = border5
        
        #调整列宽，冻结窗格
        ws.column_dimensions[get_column_letter(1)].width = 10
        ws.column_dimensions[get_column_letter(2)].width = 20
        ws.column_dimensions[get_column_letter(3)].width = 18
        ws.column_dimensions[get_column_letter(4)].width = 10            
        ws.freeze_panes = ws["E2"]
                    
    elif field == "consession":
        for i in range(2,m + 2):
            ws.row_dimensions[i].height = 15
            for j in range(1,n + 1):
                cell = ws.cell(row = i,column = j)
                cell.alignment = alignment1
                cell.font = font3
                if j <= 3:
                    cell.fill = fill2
                if (i - 1) % 4 == 0 and i != m + 1:
                    cell.border = border8
                if i == m + 1:
                    cell.border = border4
                if j == 3 or j == n:
                    cell.border = border3 
                if (j == 3 or j == n) and ((i - 1) % 4 == 0 and i != m + 1):
                    cell.border = border9

        #添加底部边框
        for j in range(1,n + 1):
            cell =ws.cell(row = m + 1,column = j)
            cell.border = border4
            if j == 3 or j == n:
                cell.border = border5
                
        #附加的汇总部分
        for i in range(0,k):
            ws.append(attach_list[i])
            ws.row_dimensions[m + 2 + i].height = 15
            for j in range(1,n + 1):
                cell = ws.cell(row = m + 2 + i,column = j)
                cell.alignment = alignment1
                cell.font = font4
                cell.number_format = "0.00"
                if j == n:
                    cell.border = border3
        
        for j in range(1,n + 1):
            cell = ws.cell(row = m + 1 + k,column = j)
            cell.border = border4
            if j == n:
                cell.border = border5

        #调整列宽，冻结窗格
        ws.column_dimensions[get_column_letter(1)].width = 10
        ws.column_dimensions[get_column_letter(2)].width = 20
        ws.column_dimensions[get_column_letter(3)].width = 15                
        ws.freeze_panes = ws["D2"]
    
    return wb