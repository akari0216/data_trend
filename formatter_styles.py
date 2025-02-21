import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


#居中对齐
alignment1 = Alignment(
    horizontal = "center",
    vertical = "center"
)

#字体1：雅黑10加粗
font1 = Font(
    name = "微软雅黑", 
    size = 10, 
    bold = True
)

#字体2：雅黑10红色
font2 = Font(
    name = "微软雅黑", 
    size = 10, 
    color = "FF0000"
)

#字体3：雅黑10
font3 = Font(
    name = "微软雅黑",
    size = 10
)

#字体4：雅黑10红色加粗
font4 = Font(
    name = "微软雅黑",
    size = 10,
    bold = True,
    color = "FF0000"
)

#淡蓝40%
fill1 = PatternFill(
    "solid",
    fgColor = "9BC2E6"
)

#淡蓝60%
fill2 = PatternFill(
    "solid",
    fgColor = "BDD7EE"
)

#边框：上下
border1 = Border(
    top = Side(border_style = "medium",color = "000000"), 
    bottom = Side(border_style = "medium",color = "000000")
)

#边框：上下右
border2 = Border(
    top = Side(border_style = "medium",color = "000000"),
    bottom = Side(border_style = "medium",color = "000000"),
    right = Side(border_style = "medium",color = "000000")
)

#边框：右
border3 = Border(
    right = Side(border_style = "medium",color = "000000")
)

#边框：下
border4 = Border(
    bottom = Side(border_style = "medium",color = "000000")
)

#边框：右下
border5 = Border(
    right = Side(border_style = "medium",color = "000000"), 
    bottom = Side(border_style = "medium",color = "000000")
)

#边框：上
border6 = Border(
    top = Side(border_style = "medium",color = "000000")
)

#边框：右上
border7 = Border(
    top = Side(border_style = "medium",color = "000000"), 
    right = Side(border_style = "medium",color = "000000")
)

#边框：细下
border8 = Border(
    bottom = Side(border_style = "thin",color = "000000")
)

#边框：细下中右
border9 = Border(
    bottom = Side(border_style = "thin",color = "000000"), 
    right = Side(border_style = "medium",color = "000000")
)

#边框：细上
border10 = Border(
    top = Side(border_style = "thin",color = "000000")
)


#边框：细上中右
border11 = Border(
    top = Side(border_style = "thin",color = "000000"), 
    right = Side(border_style = "medium",color = "000000")
)