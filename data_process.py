import pandas as pd
import numpy as np
import re
import copy
import datetime
import calendar
import openpyxl
import warnings

from database import DataBase
from sqlalchemy import text, bindparam
from collections import OrderedDict
from sql_querys import *
from excel_formatter import *


warnings.filterwarnings("ignore")


class CinemaData(object):
    def __init__(self, db_conn, op_date, excel_writer_name, cinema_code = None, combine_code = None) -> None:
        super().__init__()
        # 数据库连接，外部定义
        self.db_conn = db_conn
        # 影城编码
        self.cinema_code = cinema_code
        # 合并影城编码
        self.combine_code = combine_code
        # 自身影城信息
        self.info = []
        self.weight = [0.175]
        # 存储竞对编码的带顺序字典，键为竞对编码，值为[竞对编码，竞对名称，竞对信息，级别]
        self.compete_info_seq = OrderedDict()
        # 存储竞对编码的带顺序字典，键为竞对编码，值为权重
        self.compete_weight_seq = OrderedDict()
        # 储存所有的dataframe，即df_bo, df_people, df_occupancy, df_avg_price, df_consession的list
        self.dataframe_list = []

        self.index = None
        self.time_col = []
        # 运营日期, datetime.date类型-
        self.op_date = op_date
        self.workbook = openpyxl.Workbook()
        self.excel_writer_name = excel_writer_name
        self.excel_writer = pd.ExcelWriter(excel_writer_name)
        self.sheet_seq = OrderedDict()
    
    def time_col_completion(self, consession = False):
        if consession == False:
            self.time_col = ["专资编码", "影城名", "影城信息", "竞对级别"]
        else:
            self.time_col = ["卖品"]
        year, month, day = self.op_date.year, self.op_date.month, self.op_date.day
        for i in range(year - 2, year + 1):
            self.time_col.append(str(i)[2:] + "年")
        for i in range(1, month + 1):
            self.time_col.append(str(i) + "月")
        for i in range(1, day + 1):
            self.time_col.append(str(i) + "日")
        self.index = list(range(len(self.time_col)))

    def time_col_reset(self):
        self.time_col = []
        self.index = []

    # 得到自身影城信息
    def init_jy_info(
        self, 
        jy_info_sql, 
        combine = False, 
        cinema_code_from_combine_code_sql = None, 
        combine_name_sql = None):
        if not combine:
            jy_info = self.db_conn.sql_to_dataframe(text(jy_info_sql), params = {"cinema_code":self.cinema_code})
            for each_info in ["cinema_code", "cinema_name", "cinema_info"]:
                self.info.append(jy_info[each_info].tolist()[0])
        else:
            self.info.append(self.combine_code)
            cinema_code_list = self.cinema_code_from_combine_code(cinema_code_from_combine_code_sql)
            combine_name = self.db_conn.sql_to_dataframe(
                sql = text(combine_name_sql), 
                params = {"combine_code":self.combine_code}
            )["combine_name"].tolist()[0]
            self.info.append(combine_name)
            halls, seats = 0, 0
            for each_cinema_code in cinema_code_list:
                each_info = self.db_conn.sql_to_dataframe(text(jy_info_sql), params = {"cinema_code":each_cinema_code})
                halls += int(re.search(r"\d+厅", each_info["cinema_info"].tolist()[0]).group().replace("厅", ""))
                seats += int(re.search(r"\d+座", each_info["cinema_info"].tolist()[0]).group().replace("座", ""))
            self.info.append(str(halls) + "厅," + str(seats) + "座")
        self.info.append(None)

    # 得到竞对信息和权重的两个orderedlist
    def init_compete_info_weight_seq(self, compete_info_sql, combine = False):
        compete_info = None
        if not combine:
            compete_info = self.db_conn.sql_to_dataframe(text(compete_info_sql), params = {"cinema_code":self.cinema_code})
        else:
            compete_info = self.db_conn.sql_to_dataframe(text(compete_info_sql), params = {"cinema_code":self.combine_code})
        compete_list = compete_info["compete_cinema_code"].tolist()
        for idx in compete_info.index:
            self.compete_info_seq[compete_list[idx]] = compete_info.iloc[idx].tolist()[:-1]
            self.compete_weight_seq[compete_list[idx]] = compete_info.iloc[idx].tolist()[-1]

    def init_sheet_seq(self):
        fields = ["bo", "people", "occupancy", "avg_price"]
        sheet_names = ["票房数据", "人次数据", "上座率数据", "票价数据"]
        for each_field, each_sheet_name in zip(fields, sheet_names):
            self.sheet_seq[each_field] = each_sheet_name

    # 用于自身影院或竞对影院在某个字段上的某段时间内统计，统计方式为默认为sum，可指定为mean
    def cinema_data_op(self, dataframe, field):
        if field in ["bo", "people", "consession_amount", "consession_profit", "total_fee"]:
            return np.round(dataframe[field].sum(), 2)
        else:
            return np.round(dataframe[field].mean(), 2)

    # 对数据时间长度不够的补全0
    def check_data_completion(self, past_dataframe, field, time_field, whole_time_length):
        if len(past_dataframe[time_field].tolist()) == whole_time_length:
            return past_dataframe
        else:
            new_past_dataframe = pd.DataFrame(columns = [time_field])
            if time_field == "op_date":
                past_dataframe[time_field] = past_dataframe[time_field].apply(lambda x:x.day)
                new_past_dataframe[time_field] = np.array(list(range(1, whole_time_length + 1)))
            elif time_field == "month":
                new_past_dataframe[time_field] = np.array(list(range(1, whole_time_length + 1)))
            else:
                new_past_dataframe[time_field] = np.array(list(range(self.op_date.year - 2, self.op_date.year)))
            if past_dataframe.empty:
                new_past_dataframe[field] = 0
            else:
                new_past_dataframe = pd.merge(left = new_past_dataframe, right = past_dataframe, how = "left", on = time_field)
                new_past_dataframe.fillna(0, inplace = True)
            return new_past_dataframe

    # 指定field和cinema_code，以及data_year_sql, data_month_sql, data_daily_sql，生成数据列
    # 注意此处cinema_code既可以是自身影城的code，也可以是竞对影城的code
    def create_data_column_by_field(self, field, cinema_code, data_year_sql, data_month_sql, data_daily_sql):
        year, month = self.op_date.year, self.op_date.month
        begin_year = year - 2
        end_month = month - 1
        field_text = field
        if field == "bo":
            if cinema_code == self.cinema_code:
                field_text = "round(bo/10000, 2) as bo"
            else:
                field_text = "round(bo, 2) as bo"
        elif field == "total_fee":
            if cinema_code == self.cinema_code:
                field_text = "round(total_fee/10000, 2) as total_fee"
            else:
                field_text = "round(total_fee, 2) as total_fee"
        elif field == "people":
            field_text = "round(people/10000, 2) as people"
        elif field == "consession_amount":
            field_text = "round(consession_amount/10000, 2) as consession_amount"
        elif field == "consession_profit":
            field_text = "round(consession_profit/10000, 2) as consession_profit"

        cinema_code_text = "cinema_code" if cinema_code == self.cinema_code else "compete_cinema_code"
        past_year_data = self.db_conn.sql_to_dataframe(
            sql = text(data_year_sql.format(field_text)), 
            params = {cinema_code_text:cinema_code, "begin_year":begin_year}
        )
        past_year_data = self.check_data_completion(past_year_data, field, "year", 2)

        past_month_data = self.db_conn.sql_to_dataframe(
            sql = text(data_month_sql.format(field_text)), 
            params = {cinema_code_text:cinema_code, "year":year, "end_month":end_month}
        )
        past_month_data = self.check_data_completion(past_month_data, field, "month", end_month)
        
        if field == "bo" and cinema_code != self.cinema_code:
            field_text = "round(bo/10000, 2) as bo"
        past_daily_data = self.db_conn.sql_to_dataframe(
            sql = text(data_daily_sql.format(field_text)), 
            params = {cinema_code_text:cinema_code, "year":year, "month":month, "op_date":self.op_date}
        )
        past_daily_data = self.check_data_completion(past_daily_data, field, "op_date", self.op_date.day)

        current_month_data = self.cinema_data_op(past_daily_data, field)
        current_month_data = pd.DataFrame(data = {field:current_month_data, "month":month}, index = [0])
        current_year_data = self.cinema_data_op(
            pd.concat([past_month_data, current_month_data], 
            ignore_index = True), 
            field)
        current_year_data = pd.DataFrame(data = {field:current_year_data, "year":year}, index = [0])

        df = pd.concat(
            [past_year_data[field], current_year_data[field], past_month_data[field], current_month_data[field], past_daily_data[field]], 
            ignore_index = True)

        self.time_col_reset()
        return df

    def create_sheet_by_field(
        self, 
        field, 
        jy_data_year_sql, 
        jy_data_month_sql, 
        jy_data_daily_sql, 
        compete_data_year_sql, 
        compete_data_month_sql,
        compete_data_daily_sql):
        self.time_col_completion()

        df_sheet = pd.DataFrame(
            columns = ["time_col"], 
            data = self.time_col, 
            index = self.index
        )
        df_jy_header = pd.DataFrame(
            data = self.info, 
            index = list(range(len(self.info)))
        )
        df_jy_data = self.create_data_column_by_field(
            field = field, 
            cinema_code = self.cinema_code, 
            data_year_sql = jy_data_year_sql, 
            data_month_sql = jy_data_month_sql, 
            data_daily_sql = jy_data_daily_sql
        )
        df_jy_data = pd.concat([df_jy_header, df_jy_data], ignore_index = True)
        df_sheet = pd.concat([df_sheet, df_jy_data], axis = 1)
        for each_compete_cinema_code in self.compete_info_seq.keys():
            df_compete_header = pd.DataFrame(
                data = self.compete_info_seq[each_compete_cinema_code], 
                index = list(range(len(self.compete_info_seq[each_compete_cinema_code]))))
            df_compete_data = self.create_data_column_by_field(
                field = field, 
                cinema_code = each_compete_cinema_code, 
                data_year_sql = compete_data_year_sql, 
                data_month_sql = compete_data_month_sql, 
                data_daily_sql = compete_data_daily_sql
            )
            df_compete_data = pd.concat([df_compete_header, df_compete_data], ignore_index = True)
            df_sheet = pd.concat([df_sheet, df_compete_data], axis = 1)

        return df_sheet

    def create_sheet_by_consession(
        self, 
        jy_consession_year_sql, 
        jy_consession_month_sql, 
        jy_consession_daily_sql, 
        jy_data_year_sql, 
        jy_data_month_sql, 
        jy_data_daily_sql):
        self.time_col_completion(consession = True)
        df_sheet = pd.DataFrame(
            columns = ["time_col"], 
            data = self.time_col, 
            index = self.index
        )
        for each_field in ["consession_amount", "consession_profit"]:
            df_each_field = self.create_data_column_by_field(
                field = each_field, 
                cinema_code = self.cinema_code, 
                data_year_sql = jy_consession_year_sql, 
                data_month_sql = jy_consession_month_sql, 
                data_daily_sql = jy_consession_daily_sql)
            df_sheet[each_field] = df_each_field
        df_people = self.create_data_column_by_field(
            field = "people", 
            cinema_code = self.cinema_code, 
            data_year_sql = jy_data_year_sql, 
            data_month_sql = jy_data_month_sql, 
            data_daily_sql = jy_data_daily_sql)
        df_sheet["people"] = df_people
        df_sheet["spp"] = np.round(
            np.divide(
                df_sheet["consession_amount"], 
                df_sheet["people"], 
                out = np.zeros_like(df_sheet["consession_amount"]), 
                where = df_sheet["people"] != 0), 2
        )
        df_sheet["ppp"] = np.round(
            np.divide(
                df_sheet["consession_profit"], 
                df_sheet["people"], 
                out = np.zeros_like(df_sheet["consession_profit"]), 
                where = df_sheet["people"] != 0), 2
        )
        df_sheet.drop(columns = ["people"], inplace = True)
        length, width = df_sheet.shape[0], df_sheet.shape[1]
        df_sheet.iloc[1:, 1:] = df_sheet.iloc[:length - 1, 1:]
        df_sheet.iloc[0, 1:] = np.array(["卖品金额/万", "卖品利润/万", "spp/元", "ppp/元"])

        return df_sheet

    def create_cinema_excel(
        self, 
        jy_info_sql, 
        compete_info_sql, 
        jy_data_year_sql, 
        jy_data_month_sql, 
        jy_data_daily_sql, 
        compete_data_year_sql, 
        compete_data_month_sql,
        compete_data_daily_sql, 
        jy_market_ratio_revise_sql, 
        jy_market_ratio_target_sql, 
        jy_bo_target_sql, 
        jy_serv_fee_year_sql, 
        jy_serv_fee_month_sql, 
        jy_serv_fee_daily_sql, 
        compete_serv_fee_year_sql, 
        compete_serv_fee_month_sql, 
        compete_serv_fee_daily_sql, 
        jy_consession_year_sql, 
        jy_consession_month_sql, 
        jy_consession_daily_sql):
        self.init_sheet_seq()
        self.init_jy_info(jy_info_sql)
        self.init_compete_info_weight_seq(compete_info_sql)
        for each_field, each_sheet_name in self.sheet_seq.items():
            df_sheet = self.create_sheet_by_field(
                field = each_field, 
                jy_data_year_sql = jy_data_year_sql, 
                jy_data_month_sql = jy_data_month_sql, 
                jy_data_daily_sql = jy_data_daily_sql, 
                compete_data_year_sql = compete_data_year_sql, 
                compete_data_month_sql = compete_data_month_sql,
                compete_data_daily_sql = compete_data_daily_sql
                )
            if each_field == "bo":
                df_sheet = self.merge_serv_fee_with_bo(
                    df_sheet, 
                    jy_serv_fee_year_sql = jy_serv_fee_year_sql, 
                    jy_serv_fee_month_sql = jy_serv_fee_month_sql, 
                    jy_serv_fee_daily_sql = jy_serv_fee_daily_sql, 
                    compete_serv_fee_year_sql = compete_serv_fee_year_sql, 
                    compete_serv_fee_month_sql = compete_serv_fee_month_sql, 
                    compete_serv_fee_daily_sql = compete_serv_fee_daily_sql)
                sheet_copy = copy.deepcopy(df_sheet)
                self.dataframe_list.append(sheet_copy)
                # df_sheet = self.add_column_realtime_market_ratio(df_sheet)
                df_sheet = self.add_column_abs_market_ratio(df_sheet)
                df_sheet = self.add_column_market_ratio_revise(df_sheet, jy_market_ratio_revise_sql)
                df_sheet = self.add_column_market_ratio_target(df_sheet, jy_market_ratio_target_sql)
                df_sheet = self.add_column_achievement_rate(df_sheet)
                df_sheet = self.add_column_bo_target(df_sheet, jy_bo_target_sql)
                df_sheet = self.add_column_target_schedule(df_sheet)
                df_sheet = self.add_column_time_schedule(df_sheet)
            elif each_field == "avg_price":
                df_bo, df_people = self.dataframe_list[0], self.dataframe_list[1]
                df_sheet = copy.deepcopy(df_bo)
                length, width = df_sheet.shape[0], df_sheet.shape[1]
                for i in range(4, length):
                    for j in range(1, width):
                        if df_bo.iloc[i, j] != 0:
                            df_sheet.iloc[i, j] = np.round(
                                np.divide(
                                    df_bo.iloc[i, j], 
                                    df_people.iloc[i, j], 
                                    out = np.zeros_like(df_bo.iloc[i, j]), 
                                    where = df_people.iloc[i, j] != 0), 2)
                        else:
                            df_sheet.iloc[i, j] = 0
                
                self.dataframe_list.append(df_sheet)
            else:
                self.dataframe_list.append(df_sheet)
            sheet_copy = copy.deepcopy(df_sheet)
            self.workbook = xls_formatter_cinema(
                wb = self.workbook, 
                df = sheet_copy, 
                sheet_name = each_sheet_name, 
                weight = self.weight
            )
            # df_sheet.to_excel(
            #     excel_writer = self.excel_writer, 
            #     sheet_name = each_sheet_name, 
            #     index = False, 
            #     header = False
            # )
        df_sheet_consession = self.create_sheet_by_consession(
            jy_consession_year_sql = jy_consession_year_sql, 
            jy_consession_month_sql = jy_consession_month_sql, 
            jy_consession_daily_sql = jy_consession_daily_sql, 
            jy_data_year_sql = jy_data_year_sql, 
            jy_data_month_sql = jy_data_month_sql, 
            jy_data_daily_sql = jy_data_daily_sql)
        self.dataframe_list.append(df_sheet_consession)
        sheet_consession_copy = copy.deepcopy(df_sheet_consession)
        self.workbook = xls_formatter_cinema(
            wb = self.workbook, 
            df = sheet_consession_copy, 
            sheet_name = "卖品数据", 
            weight = self.weight
        )

        self.workbook.remove_sheet(self.workbook.get_sheet_by_name("Sheet"))
        self.workbook.save(self.excel_writer_name)
        # df_sheet_consession.to_excel(
        #     excel_writer = self.excel_writer, 
        #     sheet_name = "卖品数据", 
        #     index = False, 
        #     header = False)
        # self.excel_writer._save()

    def cinema_code_from_combine_code(self, cinema_code_from_combine_code_sql):
        cinema_code_data = self.db_conn.sql_to_dataframe(
            sql = text(cinema_code_from_combine_code_sql), 
            params = {"combine_code":self.combine_code}
        )
        return cinema_code_data["cinema_code"].tolist()

    # 合并影城的数据和卖品另外设计方法生成之
    def create_combine_cinema_sheet_by_field(
        self, 
        field, 
        cinema_code_from_combine_code_sql, 
        jy_data_year_sql, 
        jy_data_month_sql, 
        jy_data_daily_sql, 
        compete_data_year_sql, 
        compete_data_month_sql,
        compete_data_daily_sql
    ):
        self.time_col_completion()

        df_sheet = pd.DataFrame(
            columns = ["time_col"], 
            data = self.time_col, 
            index = self.index
        )
        df_jy_header = pd.DataFrame(
            data = self.info, 
            index = list(range(len(self.info)))
        )
        df_jy_data = pd.Series(data = [0] * (len(self.time_col) - 4))
        cinema_code_list = self.cinema_code_from_combine_code(cinema_code_from_combine_code_sql)
        for each_code in cinema_code_list:
            self.cinema_code = each_code
            each_cinema_data = self.create_data_column_by_field(
                field = field, 
                cinema_code = each_code, 
                data_year_sql = jy_data_year_sql, 
                data_month_sql = jy_data_month_sql, 
                data_daily_sql = jy_data_daily_sql)
            df_jy_data += each_cinema_data
        df_jy_data = pd.concat([df_jy_header, df_jy_data], ignore_index = True)
        df_sheet = pd.concat([df_sheet, df_jy_data], axis = 1)
        for each_compete_cinema_code in self.compete_info_seq.keys():
            df_compete_header = pd.DataFrame(
                data = self.compete_info_seq[each_compete_cinema_code], 
                index = list(range(len(self.compete_info_seq[each_compete_cinema_code]))))
            df_compete_data = self.create_data_column_by_field(
                field = field, 
                cinema_code = each_compete_cinema_code, 
                data_year_sql = compete_data_year_sql, 
                data_month_sql = compete_data_month_sql, 
                data_daily_sql = compete_data_daily_sql
            )
            df_compete_data = pd.concat([df_compete_header, df_compete_data], ignore_index = True)
            df_sheet = pd.concat([df_sheet, df_compete_data], axis = 1)
        
        return df_sheet

    def create_combine_cinema_excel(
        self, 
        jy_info_sql, 
        compete_info_sql, 
        cinema_code_from_combine_code_sql, 
        combine_name_sql, 
        jy_data_year_sql, 
        jy_data_month_sql, 
        jy_data_daily_sql, 
        compete_data_year_sql, 
        compete_data_month_sql,
        compete_data_daily_sql, 
        jy_market_ratio_revise_sql, 
        jy_market_ratio_target_sql, 
        jy_bo_target_sql, 
        jy_serv_fee_year_sql, 
        jy_serv_fee_month_sql, 
        jy_serv_fee_daily_sql, 
        compete_serv_fee_year_sql, 
        compete_serv_fee_month_sql, 
        compete_serv_fee_daily_sql, 
        jy_consession_year_sql, 
        jy_consession_month_sql, 
        jy_consession_daily_sql
    ):
        self.init_sheet_seq()
        self.init_jy_info(
            jy_info_sql = jy_info_sql, 
            combine = True, 
            cinema_code_from_combine_code_sql = cinema_code_from_combine_code_sql, 
            combine_name_sql = combine_name_sql
        )
        self.init_compete_info_weight_seq(compete_info_sql, combine = True)
        for each_field, each_sheet_name in self.sheet_seq.items():
            df_sheet = self.create_combine_cinema_sheet_by_field(
                field = each_field, 
                cinema_code_from_combine_code_sql = cinema_code_from_combine_code_sql, 
                jy_data_year_sql = jy_data_year_sql, 
                jy_data_month_sql = jy_data_month_sql, 
                jy_data_daily_sql = jy_data_daily_sql, 
                compete_data_year_sql = compete_data_year_sql, 
                compete_data_month_sql = compete_data_month_sql,
                compete_data_daily_sql = compete_data_daily_sql
            )
            if each_field == "bo":
                df_sheet = self.merge_serv_fee_with_bo(
                    df_sheet, 
                    jy_serv_fee_year_sql = jy_serv_fee_year_sql, 
                    jy_serv_fee_month_sql = jy_serv_fee_month_sql, 
                    jy_serv_fee_daily_sql = jy_serv_fee_daily_sql, 
                    compete_serv_fee_year_sql = compete_serv_fee_year_sql, 
                    compete_serv_fee_month_sql = compete_serv_fee_month_sql, 
                    compete_serv_fee_daily_sql = compete_serv_fee_daily_sql, 
                    combine = True, 
                    cinema_code_from_combine_code_sql = cinema_code_from_combine_code_sql)
                sheet_copy = copy.deepcopy(df_sheet)
                self.dataframe_list.append(sheet_copy)
                # df_sheet = self.add_column_realtime_market_ratio(df_sheet)
                df_sheet = self.add_column_abs_market_ratio(df_sheet)
                df_sheet = self.add_column_market_ratio_revise(df_sheet, jy_market_ratio_revise_sql, combine = True)
                df_sheet = self.add_column_market_ratio_target(df_sheet, jy_market_ratio_target_sql, combine = True)
                df_sheet = self.add_column_achievement_rate(df_sheet)
                df_sheet = self.add_column_bo_target(df_sheet, jy_bo_target_sql, combine = True)
                df_sheet = self.add_column_target_schedule(df_sheet)
                df_sheet = self.add_column_time_schedule(df_sheet)
            elif each_field == "avg_price":
                df_bo, df_people = self.dataframe_list[0], self.dataframe_list[1]
                df_sheet = copy.deepcopy(df_bo)
                length, width = df_sheet.shape[0], df_sheet.shape[1]
                for i in range(4, length):
                    for j in range(1, width):
                        df_sheet.iloc[i, j] = np.round(
                            np.divide(
                                df_bo.iloc[i, j], 
                                df_people.iloc[i, j], 
                                out = np.zeros_like(df_bo.iloc[i, j]), 
                                where = df_people.iloc[i, j] != 0), 2)
                
                self.dataframe_list.append(df_sheet)
            else:
                self.dataframe_list.append(df_sheet)
            sheet_copy = copy.deepcopy(df_sheet)
            self.workbook = xls_formatter_cinema(
                wb = self.workbook, 
                df = sheet_copy, 
                sheet_name = each_sheet_name,
                weight = self.weight
            )
            # df_sheet.to_excel(
            #     excel_writer = self.excel_writer, 
            #     sheet_name = each_sheet_name, 
            #     index = False, 
            #     header = False
            # )
        df_sheet_consession = self.create_combine_cinema_sheet_by_consession(
            cinema_code_from_combine_code_sql = cinema_code_from_combine_code_sql, 
            jy_consession_year_sql = jy_consession_year_sql, 
            jy_consession_month_sql = jy_consession_month_sql, 
            jy_consession_daily_sql = jy_consession_daily_sql, 
            jy_data_year_sql = jy_data_year_sql, 
            jy_data_month_sql = jy_data_month_sql, 
            jy_data_daily_sql = jy_data_daily_sql)
        self.dataframe_list.append(df_sheet_consession)
        sheet_consession_copy = copy.deepcopy(df_sheet_consession)
        self.workbook = xls_formatter_cinema(
            wb = self.workbook, 
            df = sheet_consession_copy, 
            sheet_name = "卖品数据",
            weight = self.weight
        )

        self.workbook.remove_sheet(self.workbook.get_sheet_by_name("Sheet"))
        self.workbook.save(self.excel_writer_name)
        # df_sheet_consession.to_excel(
        #     excel_writer = self.excel_writer, 
        #     sheet_name = "卖品数据", 
        #     index = False, 
        #     header = False)

        # self.excel_writer._save()

    def create_combine_cinema_sheet_by_consession(
        self, 
        cinema_code_from_combine_code_sql, 
        jy_consession_year_sql, 
        jy_consession_month_sql, 
        jy_consession_daily_sql, 
        jy_data_year_sql, 
        jy_data_month_sql, 
        jy_data_daily_sql):
        self.time_col_completion(consession = True)
        df_sheet = pd.DataFrame(
            columns = ["time_col"], 
            data = self.time_col, 
            index = self.index
        )
        cinema_code_list = self.cinema_code_from_combine_code(cinema_code_from_combine_code_sql)
        length = df_sheet.shape[0]
        for each_field in ["consession_amount", "consession_profit"]:
            df_each_field = pd.Series(data = [0] * length)
            for each_code in cinema_code_list:
                self.cinema_code = each_code
                df_each_code_field = self.create_data_column_by_field(
                    field = each_field, 
                    cinema_code = each_code, 
                    data_year_sql = jy_consession_year_sql, 
                    data_month_sql = jy_consession_month_sql, 
                    data_daily_sql = jy_consession_daily_sql)
                df_each_field += df_each_code_field
            df_sheet[each_field] = df_each_field
        
        df_people = pd.Series(data = [0] * length)
        for each_code in cinema_code_list:
            self.cinema_code = each_code
            df_each_code_people = self.create_data_column_by_field(
                field = "people", 
                cinema_code = each_code, 
                data_year_sql = jy_data_year_sql, 
                data_month_sql = jy_data_month_sql, 
                data_daily_sql = jy_data_daily_sql)
            df_people += df_each_code_people
        df_sheet["people"] = df_people
        df_sheet["spp"] = np.round(
            np.divide(
                df_sheet["consession_amount"], 
                df_sheet["people"], 
                out = np.zeros_like(df_sheet["consession_amount"]), 
                where = df_sheet["people"] != 0), 2
        )
        df_sheet["ppp"] = np.round(
            np.divide(
                df_sheet["consession_profit"], 
                df_sheet["people"], 
                out = np.zeros_like(df_sheet["consession_profit"]), 
                where = df_sheet["people"] != 0), 2
        )
        df_sheet.drop(columns = ["people"], inplace = True)
        length, width = df_sheet.shape[0], df_sheet.shape[1]
        df_sheet.iloc[1:, 1:] = df_sheet.iloc[:length - 1, 1:]
        df_sheet.iloc[0, 1:] = np.array(["卖品金额/万", "卖品利润/万", "spp/元", "ppp/元"])

        return df_sheet

    def merge_serv_fee_with_bo(
        self, 
        df_sheet, 
        jy_serv_fee_year_sql, 
        jy_serv_fee_month_sql, 
        jy_serv_fee_daily_sql, 
        compete_serv_fee_year_sql, 
        compete_serv_fee_month_sql, 
        compete_serv_fee_daily_sql, 
        combine = False, 
        cinema_code_from_combine_code_sql = None):
        length, width = df_sheet.shape[0], len(self.compete_weight_seq)
        df_serv_fee = pd.Series(data = [0] * (length - 4))
        if not combine:
            df_serv_fee = self.create_data_column_by_field(
                field = "total_fee", 
                cinema_code = self.cinema_code, 
                data_year_sql = jy_serv_fee_year_sql, 
                data_month_sql = jy_serv_fee_month_sql, 
                data_daily_sql = jy_serv_fee_daily_sql
            )
        else:
            cinema_code_list = self.cinema_code_from_combine_code(cinema_code_from_combine_code_sql)
            for each_code in cinema_code_list:
                self.cinema_code = each_code
                each_serv_fee = self.create_data_column_by_field(
                    field = "total_fee", 
                    cinema_code = each_code, 
                    data_year_sql = jy_serv_fee_year_sql, 
                    data_month_sql = jy_serv_fee_month_sql, 
                    data_daily_sql = jy_serv_fee_daily_sql)
                df_serv_fee += each_serv_fee
        for each_compete_cinema_code in self.compete_info_seq.keys():
            df_compete_serv_fee = self.create_data_column_by_field(
                field = "total_fee", 
                cinema_code = each_compete_cinema_code, 
                data_year_sql = compete_serv_fee_year_sql, 
                data_month_sql = compete_serv_fee_month_sql, 
                data_daily_sql = compete_serv_fee_daily_sql)
            df_serv_fee = pd.concat([df_serv_fee, df_compete_serv_fee], axis = 1)
        df_serv_fee.columns = df_sheet.columns[1:]
        df_serv_fee.index = list(range(4, length))
        df_sheet.iloc[4:length + 1, 1:width + 2] = df_sheet.iloc[4:length + 1, 1:width + 2] + df_serv_fee

        return df_sheet
    
    # 添加实时考核市占列
    def add_column_realtime_market_ratio(self, df_sheet_bo):
        length, width = df_sheet_bo.shape[0], len(self.compete_weight_seq)
        for each_compete, each_weight in self.compete_weight_seq.items():
            self.weight.append(each_weight)
        self.weight = np.array(self.weight)

        for i in range(4, length):
            df_sheet_bo.loc[i, "realtime_market_ratio"] = 0
            for j in range(1, width + 2):
                df_sheet_bo.loc[i, "realtime_market_ratio"] += df_sheet_bo.iloc[i, j] * self.weight[j - 1] 
            df_sheet_bo.loc[i, "realtime_market_ratio"] = np.round(np.divide(
                df_sheet_bo.iloc[i, 1] * self.weight[0], df_sheet_bo.loc[i, "realtime_market_ratio"]), 4)
        df_sheet_bo.loc[1, "realtime_market_ratio"] = "当日/月/年市占"
        df_sheet_bo["realtime_market_ratio"].fillna("", inplace = True)

        return df_sheet_bo        

    # 添加市占修正列
    def add_column_market_ratio_revise(self, df_sheet_bo, jy_market_ratio_revise_sql, combine = False):
        length = df_sheet_bo.shape[0]
        year, month = self.op_date.year, self.op_date.month
        df_jy_market_ratio_revise = None
        if not combine:
            df_jy_market_ratio_revise = self.db_conn.sql_to_dataframe(
                sql = text(jy_market_ratio_revise_sql), 
                params = {"cinema_code": self.cinema_code, "year": year, "month": month}
            )
        else:
            df_jy_market_ratio_revise = self.db_conn.sql_to_dataframe(
                sql = text(jy_market_ratio_revise_sql), 
                params = {"cinema_code": self.combine_code, "year": year, "month": month}
            )
        df_jy_market_ratio_revise = pd.concat(
            [pd.DataFrame(data = [0] * 7, columns = ["market_ratio_revise"]), 
            df_jy_market_ratio_revise, 
            pd.DataFrame(data = [0] * (length - df_jy_market_ratio_revise.shape[0] - 7), columns = ["market_ratio_revise"])], 
            ignore_index = True
        )
        df_sheet_bo["market_ratio_revise"] = df_jy_market_ratio_revise
        df_sheet_bo.loc[1, "market_ratio_revise"] = "实际累计市占"
        df_sheet_bo["market_ratio_revise"].fillna("", inplace = True)
        for i in [0, 2, 3]:
            df_sheet_bo.loc[i, "market_ratio_revise"] = ""

        return df_sheet_bo
    # 添加目标市占列
    def add_column_market_ratio_target(self, df_sheet_bo, jy_market_ratio_target_sql, combine = False):
        length = df_sheet_bo.shape[0]
        year, month = self.op_date.year, self.op_date.month
        df_jy_market_ratio_target = None
        if not combine:
            df_jy_market_ratio_target = self.db_conn.sql_to_dataframe(
                sql = text(jy_market_ratio_target_sql), 
                params = {"cinema_code": self.cinema_code, "year": year, "month": month}
            )
        else:
            df_jy_market_ratio_target = self.db_conn.sql_to_dataframe(
                sql = text(jy_market_ratio_target_sql), 
                params = {"cinema_code": self.combine_code, "year": year, "month": month}
            )
        df_jy_market_ratio_target = pd.concat(
            [pd.DataFrame(data = [0] * 7, columns = ["market_ratio_target"]), 
            df_jy_market_ratio_target, 
            pd.DataFrame(data = [0] * (length - df_jy_market_ratio_target.shape[0] - 7), columns = ["market_ratio_target"])], 
            ignore_index = True
        )
        df_sheet_bo["market_ratio_target"] = df_jy_market_ratio_target
        df_sheet_bo.loc[1, "market_ratio_target"] = "目标实际累计市占"
        df_sheet_bo["market_ratio_target"].fillna("", inplace = True)
        for i in [0, 2, 3]:
            df_sheet_bo.loc[i, "market_ratio_target"] = ""

        return df_sheet_bo

    # 添加达成率列
    def add_column_achievement_rate(self, df_sheet_bo):
        length = df_sheet_bo.shape[0]
        for i in range(4, length):
            df_sheet_bo.loc[i, "achievement_rate"] = np.divide(
                df_sheet_bo.loc[i, "market_ratio_revise"], 
                df_sheet_bo.loc[i, "market_ratio_target"], 
                out = np.zeros_like(df_sheet_bo.loc[i, "market_ratio_revise"], dtype = np.float64), 
                where = (df_sheet_bo.loc[i, "market_ratio_target"] != 0)
            )
        df_sheet_bo.loc[1, "achievement_rate"] = "达成率"
        df_sheet_bo["achievement_rate"].fillna("", inplace = True)

        return df_sheet_bo
    
    # 添加绝对市占列
    def add_column_abs_market_ratio(self, df_sheet_bo):
        # 由于add_column_realtime_market_ratio方法弃用，self.weight在此添加其竞对权重
        for each_compete, each_weight in self.compete_weight_seq.items():
            self.weight.append(each_weight)
        self.weight = np.array(self.weight)
        
        length, width = df_sheet_bo.shape[0], len(self.compete_weight_seq)
        for i in range(6, length):  
            df_sheet_bo.loc[i, "abs_market_ratio"] = 0
            for j in range(1, width + 2):
                df_sheet_bo.loc[i, "abs_market_ratio"] += df_sheet_bo.iloc[i, j]
            df_sheet_bo.loc[i, "abs_market_ratio"] = np.round(np.divide(
                df_sheet_bo.iloc[i, 1], df_sheet_bo.loc[i, "abs_market_ratio"]), 4)
        df_sheet_bo.loc[1, "abs_market_ratio"] = "当日/月/年绝对市占率"
        df_sheet_bo["abs_market_ratio"].fillna("", inplace = True)

        return df_sheet_bo

    # 添加票房目标列
    def add_column_bo_target(self, df_sheet_bo, jy_bo_target_sql, combine = False):
        length = df_sheet_bo.shape[0]
        year, month = self.op_date.year, self.op_date.month
        df_jy_bo_target = None
        if not combine:
            df_jy_bo_target = self.db_conn.sql_to_dataframe(
                sql = text(jy_bo_target_sql), 
                params = {"cinema_code": self.cinema_code, "year": year, "month": month}
            )
        else:
            df_jy_bo_target = self.db_conn.sql_to_dataframe(
                sql = text(jy_bo_target_sql), 
                params = {"cinema_code": self.combine_code, "year": year, "month": month}
            )
        df_jy_bo_target = pd.concat(
            [pd.DataFrame(data = [0] * 7, columns = ["bo_target_accum"]), 
            df_jy_bo_target, 
            pd.DataFrame(data = [0] * (length - df_jy_bo_target.shape[0] - 7), columns = ["bo_target_accum"])], 
            ignore_index = True
        )
        df_sheet_bo["bo_target_accum"] = df_jy_bo_target
        df_sheet_bo.loc[6, "bo_target_accum"] = df_sheet_bo.loc[6 + month, "bo_target_accum"]
        df_sheet_bo.loc[1, "bo_target_accum"] = "累计票房目标(含服务费)"
        df_sheet_bo["bo_target_accum"].fillna("", inplace = True)
        for i in [0, 2, 3]:
            df_sheet_bo.loc[i, "bo_target_accum"] = ""

        return df_sheet_bo

    # 添加目标进度列
    def add_column_target_schedule(self, df_sheet_bo):
        month = self.op_date.month
        bo_sum = 0
        for i in range(7, 7 + month):
            bo_sum += df_sheet_bo.iloc[i, 1]
            df_sheet_bo.loc[i, "target_schedule"] = np.round(np.divide(
                bo_sum, df_sheet_bo.loc[i, "bo_target_accum"]), 4)
        df_sheet_bo.loc[6, "target_schedule"] = df_sheet_bo.loc[6 + month, "target_schedule"]
        df_sheet_bo.loc[1, "target_schedule"] = "目标进度"
        df_sheet_bo["target_schedule"].fillna("", inplace = True)

        return df_sheet_bo

    # 添加时间进度列
    def add_column_time_schedule(self, df_sheet_bo):
        year, month, day = self.op_date.year, self.op_date.month, self.op_date.day
        day_sum = 0
        total_days = 365 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 366
        for i in range(7, 7 + month):
            if i != 7 + month - 1:
                day_sum += calendar.monthrange(year, i - 6)[1]
            else:
                day_sum += day
            df_sheet_bo.loc[i, "time_schedule"] = np.round(np.divide(
                day_sum, total_days), 4)
        df_sheet_bo.loc[6, "time_schedule"] = df_sheet_bo.loc[6 + month, "time_schedule"] 
        df_sheet_bo.loc[1, "time_schedule"] = "时间进度"
        df_sheet_bo["time_schedule"].fillna("", inplace = True)

        return df_sheet_bo

    def save_dataframe_list(self):
        return self.dataframe_list

    def save_weights(self):
        return self.weight






# 同城数据
class CityData(object):
    def __init__(self, db_conn, city, excel_writer_name, date) -> None:
        super().__init__()
        self.db_conn = db_conn
        self.city = city
        self.cinema_code_list = []
        self.cinema_name_list = []
        self.df_bo_list = []
        self.df_people_list = []
        self.df_occupancy_list = []
        self.df_avg_price_list = []
        self.df_consession_list = []
        self.weight_list = []
        self.sheet_seq = OrderedDict()
        self.excel_writer = pd.ExcelWriter(excel_writer_name)
        self.workbook = openpyxl.Workbook()
        self.excel_writer_name = excel_writer_name
        self.op_date = date
        self.bo_statistics = []
        self.people_statistics = []
    def init_sheet_seq(self):
        data_lists = [self.df_bo_list, self.df_people_list, self.df_occupancy_list, self.df_avg_price_list]
        sheet_names = ["票房数据", "人次数据", "上座率数据", "票价数据"]
        for each_data, each_sheet in zip(data_lists, sheet_names):
            self.sheet_seq[each_sheet] = each_data

    def create_city_sheet(self, data_list):
        df = data_list[0]
        for each_df in data_list[1:]:
            df = pd.concat([df, each_df.iloc[:, 1:]], axis = 1)
        return df

    def create_city_sheet_by_consession(self):
        for i, (each_df, each_code, each_name) in enumerate(zip(self.df_consession_list, self.cinema_code_list, self.cinema_name_list)):
            length, n = each_df.shape
            attach_df = pd.DataFrame(
                columns = each_df.columns, 
                data = np.zeros((2, n))
            )
            self.df_consession_list[i] = pd.concat([each_df, attach_df], ignore_index = True)
            self.df_consession_list[i].iloc[2:, :] = self.df_consession_list[i].iloc[:length, :]
            self.df_consession_list[i].iloc[0, :] = np.array(["专资编码"] + [each_code] * (n - 1))
            self.df_consession_list[i].iloc[1, :] = np.array(["影院名称"] + [each_name] * (n - 1))
            
        df = self.df_consession_list[0]
        for each_df in self.df_consession_list[1:]:
            df = pd.concat([df, each_df.iloc[:, 1:]], axis = 1)
        return df

    # 利用已生成的df_bo_list, df_people_list, df_occupancy_list, df_avg_price_list, df_consession_list生成excel
    def create_city_excel(
        self, 
        jy_city_market_ratio_revise_sql, 
        jy_city_market_ratio_target_sql, 
        jy_city_bo_target_sql
    ):
        self.init_sheet_seq()
        for each_sheet, each_data_list in self.sheet_seq.items():
            df_sheet = self.create_city_sheet(each_data_list)
            df_sheet = self.add_column_statistics(each_sheet, df_sheet)
            if each_sheet == "票房数据":
                # df_sheet = self.add_column_realtime_market_ratio(df_sheet)
                df_sheet = self.add_column_city_abs_market_ratio(df_sheet)
                df_sheet = self.add_column_city_market_ratio_revise(df_sheet, jy_city_market_ratio_revise_sql)
                df_sheet = self.add_column_city_market_ratio_target(df_sheet, jy_city_market_ratio_target_sql)
                df_sheet = self.add_column_city_achievement_rate(df_sheet)
                df_sheet = self.add_column_city_bo_target(df_sheet, jy_city_bo_target_sql)
                df_sheet = self.add_column_city_target_schedule(df_sheet)
                df_sheet = self.add_column_city_time_schedule(df_sheet)
            self.workbook = xls_formatter_city(
                wb = self.workbook, 
                df = df_sheet, 
                sheet_name = each_sheet, 
                weight = self.weight_list
            )
            # df_sheet.to_excel(self.excel_writer, sheet_name = each_sheet, header = False, index = False)

        df_consession = self.create_city_sheet_by_consession()
        df_consession = self.add_column_consession_statistics(df_consession)
        # df_consession.to_excel(self.excel_writer, sheet_name = "卖品数据", header = False, index = False)

        # self.excel_writer._save()
        self.workbook = xls_formatter_city(
            wb = self.workbook, 
            df = df_consession, 
            sheet_name = "卖品数据", 
            weight = self.weight_list
        )

        self.workbook.remove_sheet(self.workbook.get_sheet_by_name("Sheet"))
        self.workbook.save(self.excel_writer_name)

    def add_data_to_cinema_code_list(self, cinema_code):
        self.cinema_code_list.append(cinema_code)

    def add_data_to_cinema_name_list(self, cinema_name):
        self.cinema_name_list.append(cinema_name)

    def add_data_to_df_bo_list(self, df_bo):
        self.df_bo_list.append(df_bo)

    def add_data_to_df_people_list(self, df_people):
        self.df_people_list.append(df_people)

    def add_data_to_df_occupancy_list(self, df_occupancy):
        self.df_occupancy_list.append(df_occupancy)

    def add_data_to_df_avg_price_list(self, df_avg_price):
        self.df_avg_price_list.append(df_avg_price)

    def add_data_to_df_consession_list(self, df_consession):
        self.df_consession_list.append(df_consession)

    def add_data_to_weight_list(self, weights):
        self.weight_list.extend(weights)

    def add_column_statistics(self, df_sheet_name, df_sheet):
        length = df_sheet.shape[0]
        if df_sheet_name == "票房数据" or df_sheet_name == "人次数据":
            sheet_data = self.df_bo_list if df_sheet_name == "票房数据" else self.df_people_list
            for i in range(4, length):
                df_sheet.loc[i, "statistics"] = 0
                for each_data in sheet_data:
                    df_sheet.loc[i, "statistics"] += each_data.iloc[i, 1]
            if df_sheet_name == "票房数据":
                df_sheet.loc[1, "statistics"] = "同城票房/万"
                self.bo_statistics = df_sheet.loc[4:, "statistics"]
            else:
                df_sheet.loc[1, "statistics"] = "同城人次/万"
                self.people_statistics = df_sheet.loc[4:, "statistics"]
        elif df_sheet_name == "上座率数据":
            m = len(self.df_occupancy_list)
            for i in range(4, length):
                df_sheet.loc[i, "statistics"] = 0
                for j in range(m):
                    df_sheet.loc[i, "statistics"] += self.df_occupancy_list[j].iloc[i, 1]
                df_sheet.loc[i, "statistics"] /= m
                df_sheet.loc[i, "statistics"] = round(df_sheet.loc[i, "statistics"], 2)
            df_sheet.loc[1, "statistics"] = "同城上座率"
        else:
            for i in range(4, length):
                df_sheet.loc[i, "statistics"] = np.divide(
                    self.bo_statistics[i], 
                    self.people_statistics[i], 
                    out = np.zeros_like(self.bo_statistics[i]), 
                    where = (self.people_statistics[i] != 0)
                    )
                df_sheet.loc[i, "statistics"] = round(df_sheet.loc[i, "statistics"], 2)
            df_sheet.loc[1, "statistics"] = "同城票价/元"
        df_sheet["statistics"].fillna("", inplace = True)

        return df_sheet

    def add_column_realtime_market_ratio(self, df_sheet):
        length = df_sheet.shape[0]
        n = len(self.weight_list)
        for i in range(4, length):
            df_sheet.loc[i, "realtime_market_ratio"] = 0
            df_sheet.loc[i, "realtime_market_sum"] = 0
            for j in range(n):
                if self.weight_list[j] == 0.175:
                    df_sheet.loc[i, "realtime_market_ratio"] += self.weight_list[j] * df_sheet.iloc[i, j + 1]
                df_sheet.loc[i, "realtime_market_sum"] += self.weight_list[j] * df_sheet.iloc[i, j + 1]

            df_sheet.loc[i, "realtime_market_ratio"] = np.round(
                np.divide(
                    df_sheet.loc[i, "realtime_market_ratio"], 
                    df_sheet.loc[i, "realtime_market_sum"], 
                    out = np.zeros_like(df_sheet.loc[i, "realtime_market_ratio"]), 
                    where = (df_sheet.loc[i, "realtime_market_sum"] != 0)
                ), 4)
        df_sheet.drop(columns = ["realtime_market_sum"], axis = 1, inplace = True)
        df_sheet.loc[1, "realtime_market_ratio"] = "当日/月/年市占"
        df_sheet["realtime_market_ratio"].fillna("", inplace = True)

        return df_sheet

    def add_column_city_market_ratio_revise(self, df_sheet, jy_city_market_ratio_revise_sql):
        length = df_sheet.shape[0]
        year, month = self.op_date.year, self.op_date.month
        df_jy_city_market_ratio_revise = self.db_conn.sql_to_dataframe(
            sql = text(jy_city_market_ratio_revise_sql),  
            params = {"city": self.city, "year": year, "month": month}
        )
        df_jy_city_market_ratio_revise = pd.concat(
            [pd.DataFrame(data = [0] * 7, columns = ["market_ratio_revise"]), 
            df_jy_city_market_ratio_revise, 
            pd.DataFrame(data = [0] * (length - df_jy_city_market_ratio_revise.shape[0] - 7),
            columns = ["market_ratio_revise"])], 
            ignore_index = True
        )

        df_sheet["city_market_ratio_revise"] = df_jy_city_market_ratio_revise
        df_sheet.loc[1, "city_market_ratio_revise"] = "同城实际累计市占"
        df_sheet["city_market_ratio_revise"].fillna("", inplace = True)
        for i in [0, 2, 3]:
            df_sheet.loc[i, "city_market_ratio_revise"] = ""

        return df_sheet


    def add_column_city_market_ratio_target(self, df_sheet, jy_city_market_ratio_target_sql):
        length = df_sheet.shape[0]
        year, month = self.op_date.year, self.op_date.month
        df_jy_city_market_ratio_target = self.db_conn.sql_to_dataframe(
            sql = text(jy_city_market_ratio_target_sql),  
            params = {"city": self.city, "year": year, "month": month}
        )
        df_jy_city_market_ratio_target = pd.concat(
            [pd.DataFrame(data = [0] * 7, columns = ["market_ratio_target"]), 
            df_jy_city_market_ratio_target, 
            pd.DataFrame(data = [0] * (length - df_jy_city_market_ratio_target.shape[0] - 7),
            columns = ["market_ratio_target"])], 
            ignore_index = True
        )

        df_sheet["city_market_ratio_target"] = df_jy_city_market_ratio_target
        df_sheet.loc[1, "city_market_ratio_target"] = "同城目标市占"
        df_sheet["city_market_ratio_target"].fillna("", inplace = True)
        for i in [0, 2, 3]:
            df_sheet.loc[i, "city_market_ratio_target"] = ""

        return df_sheet

    def add_column_city_achievement_rate(self, df_sheet):
        length = df_sheet.shape[0]
        for i in range(4, length):
            df_sheet.loc[i, "city_achievement_rate"] = np.divide(
                df_sheet.loc[i, "city_market_ratio_revise"], 
                df_sheet.loc[i, "city_market_ratio_target"], 
                out = np.zeros_like(df_sheet.loc[i, "city_market_ratio_revise"]), 
                where = (df_sheet.loc[i, "city_market_ratio_target"] != 0)
            )
            df_sheet.loc[i, "city_achievement_rate"] = round(df_sheet.loc[i, "city_achievement_rate"], 4)
        df_sheet.loc[1, "city_achievement_rate"] = "同城达成率"
        df_sheet["city_achievement_rate"].fillna("", inplace = True)

        return df_sheet

    def add_column_city_abs_market_ratio(self, df_sheet):
        length = df_sheet.shape[0]
        n = len(self.weight_list)
        for i in range(6, length):
            df_sheet.loc[i, "city_abs_market_ratio"] = 0
            df_sheet.loc[i, "city_abs_market_sum"] = 0
            for j in range(n):
                if self.weight_list[j] == 0.175:
                    df_sheet.loc[i, "city_abs_market_ratio"] += df_sheet.iloc[i, j + 1]
                df_sheet.loc[i, "city_abs_market_sum"] += df_sheet.iloc[i, j + 1]
            df_sheet.loc[i, "city_abs_market_ratio"] = np.round(
                np.divide(
                    df_sheet.loc[i, "city_abs_market_ratio"], 
                    df_sheet.loc[i, "city_abs_market_sum"], 
                    out = np.zeros_like(df_sheet.loc[i, "city_abs_market_ratio"]), 
                    where = (df_sheet.loc[i, "city_abs_market_sum"] != 0),
                ), 4)

        df_sheet.drop(columns = ["city_abs_market_sum"], axis = 1, inplace = True)
        df_sheet.loc[1, "city_abs_market_ratio"] = "同城绝对市占"
        df_sheet["city_abs_market_ratio"].fillna("", inplace = True)

        return df_sheet

    def add_column_city_bo_target(self, df_sheet, jy_city_bo_target_sql):
        length = df_sheet.shape[0]
        year, month = self.op_date.year, self.op_date.month
        df_jy_city_bo_target = self.db_conn.sql_to_dataframe(
            sql = text(jy_city_bo_target_sql),  
            params = {"city": self.city, "year": year, "month": month}
        )
        df_jy_city_bo_target = pd.concat(
            [pd.DataFrame(data = [0] * 7, columns = ["bo_target_accum"]), 
            df_jy_city_bo_target, 
            pd.DataFrame(data = [0] * (length - df_jy_city_bo_target.shape[0] - 7),
            columns = ["bo_target_accum"])], 
            ignore_index = True
        )

        df_sheet["city_bo_target_accum"] = df_jy_city_bo_target
        df_sheet.loc[6, "city_bo_target_accum"] = df_sheet.loc[6 + month, "city_bo_target_accum"]
        df_sheet.loc[1, "city_bo_target_accum"] = "同城累计票房目标（含服务费）"
        df_sheet["city_bo_target_accum"].fillna("", inplace = True)
        for i in [0, 2, 3]:
            df_sheet.loc[i, "city_bo_target_accum"] = ""

        return df_sheet

    def add_column_city_target_schedule(self, df_sheet):
        month = self.op_date.month
        city_bo_sum = 0
        for i in range(7, 7 + month):
            city_bo_sum += df_sheet.loc[i, "statistics"]
            df_sheet.loc[i, "city_target_schedule"] = np.round(np.divide(
                city_bo_sum, df_sheet.loc[i, "city_bo_target_accum"]), 4)
        df_sheet.loc[6, "city_target_schedule"] = df_sheet.loc[6 + month, "city_target_schedule"]
        df_sheet.loc[1, "city_target_schedule"] = "同城进度"
        df_sheet["city_target_schedule"].fillna("", inplace = True)

        return df_sheet

    def add_column_city_time_schedule(self, df_sheet):
        year, month, day = self.op_date.year, self.op_date.month, self.op_date.day
        day_sum = 0
        total_days = 365 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 366
        for i in range(7, 7 + month):
            if i != 7 + month - 1:
                day_sum += calendar.monthrange(year, i - 6)[1]
            else:
                day_sum += day
            df_sheet.loc[i, "city_time_schedule"] = np.round(np.divide(
                day_sum, total_days), 4)
        df_sheet.loc[6, "city_time_schedule"] = df_sheet.loc[6 + month, "city_time_schedule"] 
        df_sheet.loc[1, "city_time_schedule"] = "时间进度"
        df_sheet["city_time_schedule"].fillna("", inplace = True)

        return df_sheet

    def add_column_consession_statistics(self, df_sheet):
        length = df_sheet.shape[0]
        for i in range(3, length):
            df_sheet.loc[i, "consession_amt_sum"] = 0
            df_sheet.loc[i, "consession_pft_sum"] = 0
            for each_data in self.df_consession_list:
                df_sheet.loc[i, "consession_amt_sum"] += each_data.iloc[i, 1]
                df_sheet.loc[i, "consession_pft_sum"] += each_data.iloc[i, 2]

        for i in range(3, length):
            df_sheet.loc[i, "city_spp"] = np.divide(
                df_sheet.loc[i, "consession_amt_sum"], 
                self.people_statistics[i + 1], 
                out = np.zeros_like(df_sheet.loc[i, "consession_amt_sum"]), 
                where = (self.people_statistics[i + 1] != 0)
            )
            df_sheet.loc[i, "city_spp"] = round(df_sheet.loc[i, "city_spp"], 2)

            df_sheet.loc[i, "city_ppp"] = np.divide(
                df_sheet.loc[i, "consession_pft_sum"], 
                self.people_statistics[i + 1], 
                out = np.zeros_like(df_sheet.loc[i, "consession_pft_sum"]), 
                where = (self.people_statistics[i + 1] != 0)
            )
            df_sheet.loc[i, "city_ppp"] = round(df_sheet.loc[i, "city_ppp"], 2)

        df_sheet.loc[1, "consession_amt_sum"] = "同城卖品金额/万元"
        df_sheet.loc[1, "consession_pft_sum"] = "同城卖品利润/万元"
        df_sheet.loc[1, "city_spp"] = "同城spp/元"
        df_sheet.loc[1, "city_ppp"] = "同城ppp/元"

        df_sheet["consession_amt_sum"].fillna("", inplace = True)
        df_sheet["consession_pft_sum"].fillna("", inplace = True)
        df_sheet["city_spp"].fillna("", inplace = True)
        df_sheet["city_ppp"].fillna("", inplace = True)

        return df_sheet